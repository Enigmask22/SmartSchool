"""
Subject Import/Export for Homeroom Teachers
Tính năng import môn học cho học sinh từ file Excel
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json

from core.database import get_db
from core.dependencies import get_current_user

router = APIRouter(prefix="/homeroom", tags=["homeroom-subject-import"])


@router.get("/export-subject-template/{class_name}")
async def export_subject_template(
    class_name: str,
    academic_year: Optional[str] = None,
    supabase=Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Export Excel template để nhập môn học cho học sinh
    - Môn bắt buộc (is_mandatory=true) sẽ có header đỏ và đánh dấu x mặc định
    """
    try:
        # Kiểm tra quyền giáo viên chủ nhiệm
        if current_user["role"] != "homeroom_teacher":
            raise HTTPException(status_code=403, detail="Chỉ giáo viên chủ nhiệm mới có quyền này")
        
        # Lấy danh sách học sinh của lớp
        students_query = supabase.table("students").select("*").eq("class_name", class_name).eq("is_active", True)
        
        # Thêm filter theo năm học nếu có
        if academic_year:
            # Lấy danh sách class_id từ bảng classes với năm học
            classes_response = supabase.table("classes").select("id").eq("class_name", class_name).eq("academic_year", academic_year).execute()
            if not classes_response.data:
                raise HTTPException(status_code=404, detail=f"Không tìm thấy lớp {class_name} trong năm học {academic_year}")
        
        students_response = students_query.order("student_id").execute()
        
        if not students_response.data:
            raise HTTPException(status_code=404, detail=f"Không tìm thấy học sinh nào trong lớp {class_name}")
        
        students = students_response.data
        
        # Lấy danh sách môn học đang active
        subjects_response = supabase.table("subjects").select("*").eq("is_active", True).order("is_mandatory.desc, subject_code").execute()
        
        if not subjects_response.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy môn học nào trong hệ thống")
        
        subjects = subjects_response.data
        
        # Tạo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Môn học {class_name}"
        
        # === HEADER ROW ===
        headers = ["Mã HS", "Tên học sinh"]
        subject_codes = []
        mandatory_indices = []  # Lưu index các cột môn bắt buộc
        
        for idx, subject in enumerate(subjects):
            headers.append(subject["subject_name"])
            subject_codes.append(subject["subject_code"])
            if subject["is_mandatory"]:
                mandatory_indices.append(idx + 3)  # +3 vì có Mã HS, Tên HS ở đầu (1-indexed trong Excel)
        
        # Ghi header
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Tô đỏ header môn bắt buộc
            if col_idx in mandatory_indices:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        
        # === DATA ROWS ===
        for row_idx, student in enumerate(students, start=2):
            # Mã HS
            ws.cell(row=row_idx, column=1, value=student["student_id"])
            
            # Tên HS
            ws.cell(row=row_idx, column=2, value=student["full_name"])
            
            # Lấy môn học đã chọn của học sinh (nếu có)
            selected_subjects = set()
            if student.get("subject_selected"):
                subject_data = student["subject_selected"]
                
                # Supabase JSONB trả về dict, nhưng data cũ có thể là string
                if isinstance(subject_data, str):
                    subject_data = json.loads(subject_data)
                
                # Gộp core_subjects và elective_subjects
                if isinstance(subject_data, dict):
                    if "core_subjects" in subject_data:
                        selected_subjects.update(subject_data["core_subjects"])
                    if "elective_subjects" in subject_data:
                        selected_subjects.update(subject_data["elective_subjects"])
            
            # Điền các cột môn học
            for col_idx, subject_code in enumerate(subject_codes, start=3):
                subject = subjects[col_idx - 3]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Đánh dấu x nếu:
                # 1. Môn bắt buộc (is_mandatory=True) -> x mặc định
                # 2. Học sinh đã chọn môn này trước đó
                if subject["is_mandatory"] or subject_code in selected_subjects:
                    cell.value = "x"
        
        # Định dạng cột
        ws.column_dimensions["A"].width = 15  # Mã HS
        ws.column_dimensions["B"].width = 30  # Tên HS
        for col_idx in range(3, 3 + len(subjects)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # === GHI CHÚ SHEET ===
        note_row = len(students) + 3
        ws.cell(row=note_row, column=1, value="Ghi chú:").font = Font(bold=True, color="FF0000")
        ws.cell(row=note_row + 1, column=1, value="• Các môn có header màu ĐỎ là môn BẮT BUỘC")
        ws.cell(row=note_row + 2, column=1, value="• Điền chữ 'x' vào ô nếu học sinh chọn môn đó")
        ws.cell(row=note_row + 3, column=1, value="• Để trống nếu học sinh KHÔNG chọn môn đó")
        ws.cell(row=note_row + 4, column=1, value="• KHÔNG được xóa hoặc sửa cột Mã HS và Tên học sinh")
        
        # Lưu vào BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Trả về file
        filename = f"Mau_nhap_mon_hoc_{class_name}_{academic_year or 'all'}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi tạo file mẫu: {str(e)}")


@router.post("/import-subjects/{class_name}")
async def import_subjects(
    class_name: str,
    file: UploadFile = File(...),
    academic_year: Optional[str] = None,
    supabase=Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Import môn học cho học sinh từ file Excel
    """
    try:
        # Kiểm tra quyền
        if current_user["role"] != "homeroom_teacher":
            raise HTTPException(status_code=403, detail="Chỉ giáo viên chủ nhiệm mới có quyền này")
        
        # Kiểm tra file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File phải có định dạng Excel (.xlsx hoặc .xls)")
        
        # Đọc file Excel
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Đọc header (row 1)
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(cell_value)
        
        # Validate header
        if len(headers) < 3 or headers[0] != "Mã HS" or headers[1] != "Tên học sinh":
            raise HTTPException(
                status_code=400, 
                detail="File không đúng định dạng. Header phải bắt đầu bằng 'Mã HS' và 'Tên học sinh'"
            )
        
        # Lấy danh sách môn học từ header
        subject_names = headers[2:]  # Bỏ Mã HS và Tên HS
        
        # Map subject_name -> subject_code và is_mandatory
        subjects_response = supabase.table("subjects").select("*").eq("is_active", True).execute()
        subject_map = {}
        mandatory_subjects = set()
        
        for subject in subjects_response.data:
            subject_map[subject["subject_name"]] = subject["subject_code"]
            if subject["is_mandatory"]:
                mandatory_subjects.add(subject["subject_code"])
        
        # Validate subjects trong file
        for subject_name in subject_names:
            if subject_name not in subject_map:
                raise HTTPException(
                    status_code=400,
                    detail=f"Môn học '{subject_name}' không tồn tại trong hệ thống"
                )
        
        # Đọc data và cập nhật
        updates = []
        errors = []
        
        for row_idx in range(2, ws.max_row + 1):
            student_id_cell = ws.cell(row=row_idx, column=1).value
            student_name_cell = ws.cell(row=row_idx, column=2).value
            
            # Skip empty rows
            if not student_id_cell:
                continue
            
            student_id = str(student_id_cell).strip()
            
            # Đọc các môn học được chọn
            selected_core = []
            selected_elective = []
            
            for col_idx, subject_name in enumerate(subject_names, start=3):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                
                # Nếu có "x" hoặc "X" thì là đã chọn
                if cell_value and str(cell_value).strip().lower() == 'x':
                    subject_code = subject_map[subject_name]
                    
                    # Phân loại môn bắt buộc vs tự chọn
                    if subject_code in mandatory_subjects:
                        selected_core.append(subject_code)
                    else:
                        selected_elective.append(subject_code)
            
            # Tạo JSON subject_selected
            subject_selected = {
                "core_subjects": selected_core,
                "elective_subjects": selected_elective
            }
            
            # Cập nhật database
            # Lưu ý: Supabase JSONB column tự động xử lý dict, không cần json.dumps()
            try:
                update_response = supabase.table("students").update({
                    "subject_selected": subject_selected
                }).eq("student_id", student_id).eq("class_name", class_name).execute()
                
                if not update_response.data:
                    errors.append(f"Không tìm thấy học sinh {student_id} - {student_name_cell}")
                else:
                    updates.append(student_id)
                    
            except Exception as e:
                errors.append(f"Lỗi khi cập nhật {student_id}: {str(e)}")
        
        return {
            "success": True,
            "message": f"Import thành công {len(updates)} học sinh",
            "total_updated": len(updates),
            "total_errors": len(errors),
            "updated_students": updates,
            "errors": errors if errors else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi import file: {str(e)}")

